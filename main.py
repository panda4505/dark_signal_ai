from fastapi import FastAPI, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse
from pydantic import BaseModel
from openai import OpenAI
from dotenv import load_dotenv
from openpyxl import load_workbook
import base64
import csv
import fitz
import json
import logging
import mimetypes
import os, tempfile
import re
import time
from uuid import uuid4
import zipfile
import xml.etree.ElementTree as ET

load_dotenv()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()

TEXT_MODEL = os.getenv("DEEPSEEK_MODEL", "deepseek-reasoner")
VISION_MODEL = os.getenv("DEEPSEEK_VISION_MODEL", "gpt-4o")
API_TIMEOUT_SECONDS = 900

client = OpenAI(
    api_key=os.getenv("DEEPSEEK_API_KEY"),
    base_url="https://api.deepseek.com",
    timeout=API_TIMEOUT_SECONDS,
)
openai_client = None
if os.getenv("OPENAI_API_KEY"):
    openai_client = OpenAI(
        api_key=os.getenv("OPENAI_API_KEY"),
        base_url="https://api.openai.com/v1"
    )

# Temp directory for uploaded files
UPLOAD_DIR = os.path.join(tempfile.gettempdir(), "ai_app_uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
UPLOAD_METADATA_SUFFIX = ".meta.json"
UPLOAD_CHUNK_SIZE = 1024 * 1024
MAX_UPLOAD_BYTES = 25 * 1024 * 1024
UPLOAD_TTL_SECONDS = 24 * 60 * 60
CLEANUP_INTERVAL_SECONDS = 300
LAST_CLEANUP_AT = 0.0
FILE_ID_PATTERN = re.compile(r"^[0-9a-f]{32}$")
MAX_PROMPT_CHARS = 10000


class UploadTooLargeError(ValueError):
    pass


def sanitize_uploaded_filename(filename: str | None) -> str:
    name = (filename or "upload").replace("\\", "/").rsplit("/", 1)[-1]
    return name or "upload"


def get_upload_metadata_path(file_id: str) -> str:
    return os.path.join(UPLOAD_DIR, f"{file_id}{UPLOAD_METADATA_SUFFIX}")


def safe_remove_file(path: str) -> None:
    try:
        os.remove(path)
    except FileNotFoundError:
        pass
    except OSError:
        logger.warning("Could not remove file during cleanup: %s", path, exc_info=True)


def remove_upload_artifacts(file_id: str, metadata: dict | None = None) -> None:
    safe_remove_file(get_upload_metadata_path(file_id))

    stored_name = metadata.get("stored_name") if metadata else None
    if stored_name:
        safe_remove_file(os.path.join(UPLOAD_DIR, stored_name))


def cleanup_corrupt_upload_artifacts(file_id: str) -> None:
    safe_remove_file(get_upload_metadata_path(file_id))

    try:
        entries = list(os.scandir(UPLOAD_DIR))
    except FileNotFoundError:
        return
    except OSError:
        logger.warning("Could not scan upload directory while cleaning corrupt metadata.", exc_info=True)
        return

    file_prefix = f"{file_id}."

    for entry in entries:
        if not entry.is_file():
            continue
        if entry.name != file_id and not entry.name.startswith(file_prefix):
            continue
        safe_remove_file(entry.path)


def load_upload_metadata(file_id: str) -> dict | None:
    metadata_path = get_upload_metadata_path(file_id)
    if not os.path.isfile(metadata_path):
        return None

    try:
        with open(metadata_path, "r", encoding="utf-8") as f:
            metadata = json.load(f)
    except (OSError, json.JSONDecodeError):
        logger.warning("Unreadable upload metadata for file_id=%s. Cleaning related artifacts.", file_id)
        cleanup_corrupt_upload_artifacts(file_id)
        return None

    if not isinstance(metadata, dict):
        logger.warning("Invalid upload metadata structure for file_id=%s. Cleaning related artifacts.", file_id)
        cleanup_corrupt_upload_artifacts(file_id)
        return None

    return metadata


def cleanup_expired_uploads() -> None:
    now = time.time()
    cutoff = now - UPLOAD_TTL_SECONDS

    try:
        entries = list(os.scandir(UPLOAD_DIR))
    except FileNotFoundError:
        return

    for entry in entries:
        if not entry.is_file() or not entry.name.endswith(UPLOAD_METADATA_SUFFIX):
            continue

        file_id = entry.name[:-len(UPLOAD_METADATA_SUFFIX)]
        if not FILE_ID_PATTERN.fullmatch(file_id):
            continue

        metadata = load_upload_metadata(file_id)
        created_at = 0.0
        if metadata:
            try:
                created_at = float(metadata.get("created_at", 0))
            except (TypeError, ValueError):
                created_at = 0.0

        if created_at and created_at >= cutoff:
            continue

        remove_upload_artifacts(file_id, metadata)

    for entry in entries:
        if not entry.is_file() or entry.name.endswith(UPLOAD_METADATA_SUFFIX):
            continue

        try:
            is_expired = entry.stat().st_mtime < cutoff
        except FileNotFoundError:
            continue

        if is_expired:
            safe_remove_file(entry.path)


def maybe_cleanup_expired_uploads(force: bool = False) -> None:
    global LAST_CLEANUP_AT

    now = time.time()
    if not force and LAST_CLEANUP_AT and (now - LAST_CLEANUP_AT) < CLEANUP_INTERVAL_SECONDS:
        return

    cleanup_expired_uploads()
    LAST_CLEANUP_AT = now


@app.on_event("startup")
def startup_cleanup() -> None:
    maybe_cleanup_expired_uploads(force=True)


# --- Upload endpoint: just saves the file ---

@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    maybe_cleanup_expired_uploads()

    original_name = sanitize_uploaded_filename(file.filename)
    _, ext = os.path.splitext(original_name)
    file_id = uuid4().hex
    stored_name = f"{file_id}{ext.lower()}"
    path = os.path.join(UPLOAD_DIR, stored_name)
    metadata = {
        "file_id": file_id,
        "stored_name": stored_name,
        "original_name": original_name,
        "created_at": time.time(),
    }
    bytes_written = 0

    try:
        with open(path, "wb") as f:
            while True:
                chunk = await file.read(UPLOAD_CHUNK_SIZE)
                if not chunk:
                    break

                bytes_written += len(chunk)
                if bytes_written > MAX_UPLOAD_BYTES:
                    raise UploadTooLargeError(
                        f"Files larger than {MAX_UPLOAD_BYTES // (1024 * 1024)} MB are not supported."
                    )

                f.write(chunk)

        with open(get_upload_metadata_path(file_id), "w", encoding="utf-8") as f:
            json.dump(metadata, f)
    except UploadTooLargeError as e:
        remove_upload_artifacts(file_id, metadata)
        return JSONResponse(status_code=413, content={"error": str(e)})
    except Exception:
        logger.exception("Unexpected error while saving uploaded file.")
        remove_upload_artifacts(file_id, metadata)
        return JSONResponse(
            status_code=500,
            content={"error": "Could not save the uploaded file."},
        )
    finally:
        await file.close()

    return {"file_id": file_id, "filename": original_name}


# --- Ask endpoint ---

class PromptRequest(BaseModel):
    prompt: str
    file_id: str | None = None


TEXT_FILE_EXTENSIONS = {
    "c", "cfg", "conf", "cpp", "css", "csv", "eml", "go", "html", "htm", "ini",
    "java", "js", "json", "log", "md", "py", "rb", "rs", "rtf", "sh", "sql",
    "tex", "toml", "ts", "tsx", "txt", "xml", "yaml", "yml",
}
SPREADSHEET_EXTENSIONS = {"xlsx", "xlsm", "xltx", "xltm"}
WORD_EXTENSIONS = {"docx", "docm"}
PRESENTATION_EXTENSIONS = {"pptx", "pptm"}
OPENDOCUMENT_EXTENSIONS = {"odt", "ods", "odp"}
IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "bmp", "gif", "tif", "tiff"}
LEGACY_OFFICE_EXTENSIONS = {"doc", "xls", "ppt"}
MAX_EXTRACTED_CHARS = 50000
MAX_IMAGE_BYTES = 8 * 1024 * 1024
NEWLINE = chr(10)
LARGE_TABLE_ROW_THRESHOLD = 500
TABLE_SAMPLE_HEAD_ROWS = 10
TABLE_SAMPLE_TAIL_ROWS = 5
COLUMN_SAMPLE_VALUE_LIMIT = 5


def limit_text(text: str) -> str:
    text = text.strip()
    if len(text) <= MAX_EXTRACTED_CHARS:
        return text

    suffix = "\n\n[Truncated]"
    cutoff = max(0, MAX_EXTRACTED_CHARS - len(suffix))
    return text[:cutoff].rstrip() + suffix


def get_extension(file_path: str, file_name: str | None) -> str:
    name = file_name or file_path
    return name.rsplit(".", 1)[-1].lower() if "." in name else ""


def with_fallback_text(text: str, fallback: str) -> str:
    return limit_text(text) or fallback


def normalize_tabular_row(values) -> list[str]:
    row = []

    for value in values:
        row.append("" if value is None else str(value))

    while row and not row[-1].strip():
        row.pop()

    return row


def row_has_content(row: list[str]) -> bool:
    return any(value.strip() for value in row)


def format_tabular_row(row: list[str]) -> str:
    return ",".join(row).rstrip(",")


def get_column_name(header_row: list[str], index: int) -> str:
    if index < len(header_row) and header_row[index].strip():
        return header_row[index].strip()
    return "Column " + str(index + 1)


def collect_sample_values(values: list[str]) -> list[str]:
    samples = []
    seen = set()

    for value in values:
        if value in seen:
            continue
        seen.add(value)
        samples.append(value)
        if len(samples) >= COLUMN_SAMPLE_VALUE_LIMIT:
            break

    return samples


def parse_numeric_value(value: str) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def format_numeric_value(value: float) -> str:
    return f"{value:g}"


def build_column_statistics_line(column_name: str, values: list[str]) -> str:
    unique_values = len(set(values))
    sample_values = collect_sample_values(values)
    line = (
        "- "
        + column_name
        + ": non-empty="
        + str(len(values))
        + ", unique="
        + str(unique_values)
        + ", samples="
        + json.dumps(sample_values, ensure_ascii=False)
    )

    if values:
        numeric_values = []
        is_numeric_column = True

        for value in values:
            numeric_value = parse_numeric_value(value)
            if numeric_value is None:
                is_numeric_column = False
                break
            numeric_values.append(numeric_value)

        if is_numeric_column and numeric_values:
            average = sum(numeric_values) / len(numeric_values)
            line = (
                line
                + ", min="
                + format_numeric_value(min(numeric_values))
                + ", max="
                + format_numeric_value(max(numeric_values))
                + ", average="
                + format_numeric_value(average)
            )

    return line


def build_table_statistics_block(display_label: str, rows: list[list[str]]) -> str:
    column_count = max((len(row) for row in rows), default=0)
    lines = [
        display_label,
        "Rows: " + str(len(rows)),
        "Columns: " + str(column_count),
    ]

    if not rows or column_count == 0:
        lines.append("Column names: [No columns found]")
        lines.append("Column statistics:")
        lines.append("- No data rows available.")
        return NEWLINE.join(lines)

    header_row = rows[0]
    column_names = [get_column_name(header_row, index) for index in range(column_count)]
    lines.append("Column names: " + ", ".join(column_names))
    lines.append("Column statistics:")

    data_rows = rows[1:]
    if not data_rows:
        lines.append("- No data rows available.")
        return NEWLINE.join(lines)

    for index, column_name in enumerate(column_names):
        values = []
        for row in data_rows:
            value = row[index].strip() if index < len(row) else ""
            if value:
                values.append(value)

        lines.append(build_column_statistics_line(column_name, values))

    return NEWLINE.join(lines)


def build_raw_sample_sections(tables: list[dict]) -> str:
    all_rows = []

    for table in tables:
        raw_prefix = table.get("raw_prefix", "")
        for row in table["rows"]:
            row_text = format_tabular_row(row)
            all_rows.append(raw_prefix + row_text if raw_prefix else row_text)

    if not all_rows:
        return ""

    lines = ["First 10 rows (raw text):"]
    lines.extend(all_rows[:TABLE_SAMPLE_HEAD_ROWS])
    lines.append("")
    lines.append("Last 5 rows (raw text):")
    lines.extend(all_rows[-TABLE_SAMPLE_TAIL_ROWS:])
    return NEWLINE.join(lines)


def build_large_table_analysis(tables: list[dict]) -> str:
    total_rows = sum(len(table["rows"]) for table in tables)
    total_columns = max(
        (max((len(row) for row in table["rows"]), default=0) for table in tables),
        default=0,
    )
    summary = (
        "=== LARGE FILE ANALYSIS ==="
        + NEWLINE
        + "This file has "
        + str(total_rows)
        + " rows and "
        + str(total_columns)
        + " columns. Full raw data cannot be shown. Below is a statistical summary and sample data."
        + NEWLINE
        + NEWLINE
    )
    statistics_blocks = [
        build_table_statistics_block(table["display_label"], table["rows"])
        for table in tables
    ]
    raw_sample_sections = build_raw_sample_sections(tables)

    if statistics_blocks:
        summary = summary + (NEWLINE + NEWLINE).join(statistics_blocks)

    if raw_sample_sections:
        if statistics_blocks:
            summary = summary + NEWLINE + NEWLINE
        summary = summary + raw_sample_sections

    return summary


def natural_sort_key(text: str):
    parts = []
    chunk = ""
    chunk_is_digit = None

    for char in text:
        is_digit = char.isdigit()
        if chunk_is_digit is None or is_digit == chunk_is_digit:
            chunk += char
        else:
            parts.append(int(chunk) if chunk_is_digit else chunk.lower())
            chunk = char
        chunk_is_digit = is_digit

    if chunk:
        parts.append(int(chunk) if chunk_is_digit else chunk.lower())

    return parts


def collect_xml_text(element: ET.Element) -> str:
    parts = []

    for node in element.iter():
        if node.tag.endswith(("}tab", "}br", "}cr", "}line-break")):
            parts.append(" ")
        if node.text:
            parts.append(node.text)

    return " ".join("".join(parts).split())


def extract_pdf_text(file_path: str) -> str:
    with fitz.open(file_path) as doc:
        text = "\n".join(page.get_text() for page in doc)
    return with_fallback_text(text, "[No extractable text found in PDF]")


def extract_plain_text(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        text = f.read()

    if get_extension(file_path, None) != "csv":
        return with_fallback_text(text, "[No extractable text found in file]")

    rows = []
    with open(file_path, "r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            normalized_row = normalize_tabular_row(row)
            if row_has_content(normalized_row):
                rows.append(normalized_row)

    if len(rows) > LARGE_TABLE_ROW_THRESHOLD:
        summary = build_large_table_analysis(
            [
                {
                    "display_label": "[CSV]",
                    "raw_prefix": "",
                    "rows": rows,
                }
            ]
        )
        return with_fallback_text(summary, "[No extractable text found in file]")

    return with_fallback_text(text, "[No extractable text found in file]")


def extract_spreadsheet_text(file_path: str) -> str:
    workbook = load_workbook(file_path, data_only=True)
    try:
        sections = []
        tables = []

        for sheet in workbook.worksheets:
            rows = []
            raw_rows = []
            for row in sheet.iter_rows(values_only=True):
                values = normalize_tabular_row(row)
                if row_has_content(values):
                    rows.append(values)
                    raw_rows.append(format_tabular_row(values))
            sheet_text = NEWLINE.join(raw_rows).strip()
            if sheet_text:
                sections.append("[Sheet: " + sheet.title + "]" + NEWLINE + sheet_text)
                tables.append(
                    {
                        "display_label": "[Sheet: " + sheet.title + "]",
                        "raw_prefix": "[Sheet: " + sheet.title + "] ",
                        "rows": rows,
                    }
                )

        total_rows = sum(len(table["rows"]) for table in tables)
        if total_rows > LARGE_TABLE_ROW_THRESHOLD:
            summary = build_large_table_analysis(tables)
            return with_fallback_text(
                summary,
                "[No extractable text found in spreadsheet]",
            )

        return with_fallback_text(
            (NEWLINE + NEWLINE).join(sections),
            "[No extractable text found in spreadsheet]",
        )
    finally:
        workbook.close()


def extract_docx_text(file_path: str) -> str:
    doc_members = []

    with zipfile.ZipFile(file_path) as archive:
        for name in archive.namelist():
            if name == "word/document.xml":
                doc_members.append(name)
            elif name.startswith("word/header") and name.endswith(".xml"):
                doc_members.append(name)
            elif name.startswith("word/footer") and name.endswith(".xml"):
                doc_members.append(name)
            elif name in {"word/comments.xml", "word/footnotes.xml", "word/endnotes.xml"}:
                doc_members.append(name)

        sections = []
        for name in sorted(doc_members, key=natural_sort_key):
            root = ET.fromstring(archive.read(name))
            paragraphs = []
            for element in root.iter():
                if element.tag.endswith("}p"):
                    text = collect_xml_text(element)
                    if text:
                        paragraphs.append(text)

            if paragraphs:
                label = "Document" if name == "word/document.xml" else name.rsplit("/", 1)[-1].replace(".xml", "")
                sections.append(f"[{label}]\n" + "\n".join(paragraphs))

    return with_fallback_text("\n\n".join(sections), "[No extractable text found in document]")


def extract_presentation_text(file_path: str) -> str:
    with zipfile.ZipFile(file_path) as archive:
        slide_members = sorted(
            [
                name
                for name in archive.namelist()
                if name.startswith("ppt/slides/slide") and name.endswith(".xml")
            ],
            key=natural_sort_key,
        )

        sections = []
        for index, name in enumerate(slide_members, start=1):
            root = ET.fromstring(archive.read(name))
            texts = []
            for element in root.iter():
                if element.tag.endswith("}t") and element.text:
                    chunk = " ".join(element.text.split())
                    if chunk:
                        texts.append(chunk)

            if texts:
                sections.append(f"[Slide {index}]\n" + "\n".join(texts))

    return with_fallback_text("\n\n".join(sections), "[No extractable text found in presentation]")


def extract_opendocument_text(file_path: str) -> str:
    with zipfile.ZipFile(file_path) as archive:
        if "content.xml" not in archive.namelist():
            return "[No extractable text found in document]"

        root = ET.fromstring(archive.read("content.xml"))
        blocks = []

        for element in root.iter():
            if element.tag.endswith(("}p", "}h", "}list-item", "}table-cell")):
                text = collect_xml_text(element)
                if text:
                    blocks.append(text)

    return with_fallback_text("\n".join(blocks), "[No extractable text found in document]")


def build_image_data_url(file_path: str, file_name: str | None) -> str:
    with open(file_path, "rb") as f:
        data = f.read()

    if len(data) > MAX_IMAGE_BYTES:
        raise ValueError("Image files larger than 8 MB are not supported yet.")

    mime_type = mimetypes.guess_type(file_name or file_path)[0] or "application/octet-stream"
    encoded = base64.b64encode(data).decode("ascii")
    return f"data:{mime_type};base64,{encoded}"


def build_file_context(file_path: str, file_name: str | None) -> dict:
    ext = get_extension(file_path, file_name)

    if ext in LEGACY_OFFICE_EXTENSIONS:
        return {
            "kind": "text",
            "content": (
                f"[Unsupported legacy Office file type: .{ext}. Please re-save the file as a modern format "
                "such as .docx, .xlsx, or .pptx.]"
            ),
        }

    try:
        if ext == "pdf":
            return {"kind": "text", "content": extract_pdf_text(file_path)}
        if ext in TEXT_FILE_EXTENSIONS:
            return {"kind": "text", "content": extract_plain_text(file_path)}
        if ext in SPREADSHEET_EXTENSIONS:
            return {"kind": "text", "content": extract_spreadsheet_text(file_path)}
        if ext in WORD_EXTENSIONS:
            return {"kind": "text", "content": extract_docx_text(file_path)}
        if ext in PRESENTATION_EXTENSIONS:
            return {"kind": "text", "content": extract_presentation_text(file_path)}
        if ext in OPENDOCUMENT_EXTENSIONS:
            return {"kind": "text", "content": extract_opendocument_text(file_path)}
        if ext in IMAGE_EXTENSIONS:
            return {"kind": "image", "content": build_image_data_url(file_path, file_name)}
        return {"kind": "text", "content": "[Unsupported file type]"}
    except Exception:
        logger.exception("Could not extract uploaded file content.")
        return {
            "kind": "text",
            "content": "[Could not extract readable content from the uploaded file.]",
        }


def resolve_uploaded_file(file_id: str | None):
    maybe_cleanup_expired_uploads()

    if not file_id:
        return None

    if not FILE_ID_PATTERN.fullmatch(file_id):
        raise ValueError("Invalid uploaded file reference. Please upload the file again.")

    metadata = load_upload_metadata(file_id)
    if not metadata:
        raise ValueError("Uploaded file was not found. Please upload it again.")

    try:
        created_at = float(metadata.get("created_at", 0))
    except (TypeError, ValueError):
        created_at = 0.0

    if created_at and (time.time() - created_at) > UPLOAD_TTL_SECONDS:
        remove_upload_artifacts(file_id, metadata)
        raise ValueError("Uploaded file has expired. Please upload it again.")

    stored_name = metadata.get("stored_name")
    if not stored_name:
        cleanup_corrupt_upload_artifacts(file_id)
        raise ValueError("Uploaded file metadata is invalid. Please upload it again.")

    file_path = os.path.join(UPLOAD_DIR, stored_name)
    if not os.path.isfile(file_path):
        remove_upload_artifacts(file_id, metadata)
        raise ValueError("Uploaded file is no longer available. Please upload it again.")

    return file_path, metadata.get("original_name") or "upload"


def build_request_payload(data: PromptRequest):
    if len(data.prompt) > MAX_PROMPT_CHARS:
        raise ValueError(f"Prompt is too long. Please keep it under {MAX_PROMPT_CHARS} characters.")

    uploaded_file = resolve_uploaded_file(data.file_id)

    if not uploaded_file:
        return "none", TEXT_MODEL, [
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": data.prompt},
        ]

    file_path, file_name = uploaded_file
    file_context = build_file_context(file_path, file_name)

    if file_context["kind"] == "image":
        return "image", VISION_MODEL, [
            {
                "role": "system",
                "content": (
                    "You are a helpful assistant. Analyze the uploaded image and answer the user's question. "
                    "Be honest if the image is blurry or does not contain enough information."
                ),
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": (
                            f"Filename: {file_name or '[Uploaded image]'}\n"
                            f"User question: {data.prompt}"
                        ),
                    },
                    {
                        "type": "image_url",
                        "image_url": {"url": file_context["content"]},
                    },
                ],
            },
        ]

    return "text", TEXT_MODEL, [
        {
            "role": "system",
            "content": (
                "You are a helpful assistant. Answer the user's question using the provided file content when it "
                "is available. If the extracted content says the file is unsupported or no text could be found, "
                "explain that clearly and suggest a better format when appropriate."
            ),
        },
        {
            "role": "user",
            "content": (
                f"User question:\n{data.prompt}\n\n"
                f"Filename: {file_name or '[No file uploaded]'}\n\n"
                f"Extracted file content:\n---\n{file_context['content']}\n---"
            ),
        },
    ]


@app.post("/ask")
def ask_ai(data: PromptRequest):
    file_kind = "none"
    try:
        file_kind, model_name, messages = build_request_payload(data)
        if file_kind == "image":
            if openai_client is None:
                return JSONResponse(
                    status_code=500,
                    content={
                        "error": "OPENAI_API_KEY is not configured. Please add it to your environment variables for image analysis."
                    },
                )
            response = openai_client.chat.completions.create(
                model=model_name,
                messages=messages,
                max_tokens=4096,
            )
        else:
            response = client.chat.completions.create(
                model=model_name,
                messages=messages,
                max_tokens=16384,
            )
        reasoning = getattr(response.choices[0].message, "reasoning_content", None) or ""
        content = response.choices[0].message.content or ""
        if reasoning:
            return {"answer": content, "reasoning": reasoning}
        return {"answer": content}
    except ValueError as e:
        return JSONResponse(status_code=400, content={"error": str(e)})
    except Exception:
        if file_kind == "image":
            logger.exception("Image analysis failed.")
            return JSONResponse(
                status_code=500,
                content={
                    "error": (
                        "Image analysis failed with the configured DeepSeek model. "
                        "If your current model is text-only, set DEEPSEEK_VISION_MODEL to a vision-capable model."
                    )
                },
            )
        logger.exception("Unexpected error while processing ask request.")
        return JSONResponse(
            status_code=500,
            content={"error": "Something went wrong while processing your request."},
        )


# --- Frontend ---

@app.get("/", response_class=HTMLResponse)
def home():
    return """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Dark Signal AI</title>
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <style>
            @import url("https://fonts.googleapis.com/css2?family=Outfit:wght@100..900&family=Inter:wght@400;500;600;700&family=Space+Mono:wght@400;700&display=swap");

            :root {
                color-scheme: dark;
                --font-sans: system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
                --font-mono: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
                --page-bg: #09070d;
                --panel-bg: #14101c;
                --panel-bg-alt: #181320;
                --field-bg: #100c16;
                --field-bg-active: #120f19;
                --console-bg: #0d0913;
                --border: #2b2237;
                --border-strong: #352a44;
                --accent: #8b5cf6;
                --accent-hover: #9d72ff;
                --accent-muted: #1d1530;
                --busy: #7f8bff;
                --lime: #a3e635;
                --text-primary: #f0ecf8;
                --text-secondary: #b4aac5;
                --text-muted: #887d99;
                --placeholder: #6f657f;
            }

            * {
                box-sizing: border-box;
                margin: 0;
                padding: 0;
            }

            html {
                background-color: var(--page-bg);
            }

            body {
                min-height: 100vh;
                background-color: var(--page-bg);
                color: var(--text-primary);
                font-family: var(--font-sans);
            }

            .page-shell {
                width: min(1600px, calc(100% - 40px));
                margin: 0 auto;
                padding: 28px 0 36px;
                display: grid;
                gap: 18px;
            }

            .panel {
                border: 1px solid var(--border);
                border-radius: 14px;
                background-color: var(--panel-bg);
            }

            .top-bar {
                padding: 14px 18px;
                background-color: var(--panel-bg-alt);
            }

            .topbar-row {
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 16px;
                min-height: 34px;
            }

            .brand {
                display: flex;
                align-items: center;
                gap: 10px;
                min-width: 0;
            }

            .brand-icon {
                width: 20px;
                height: 20px;
                color: var(--accent);
                flex-shrink: 0;
            }

            .app-title {
                color: var(--text-primary);
                font-family: "Outfit", sans-serif;
                font-size: 26px;
                font-weight: 700;
                letter-spacing: 0.06em;
                text-transform: uppercase;
                white-space: nowrap;
            }

            .topbar-pills {
                display: flex;
                align-items: center;
                justify-content: flex-end;
                gap: 8px;
                min-width: 0;
                flex: 1;
                flex-wrap: wrap;
            }

            .content-grid {
                display: grid;
                gap: 18px;
                align-items: start;
            }

            @media (min-width: 900px) {
                .content-grid {
                    grid-template-columns: minmax(0, 40fr) minmax(0, 60fr);
                }
            }

            .composer,
            .output-panel {
                padding: 20px;
            }

            .output-panel {
                background-color: var(--panel-bg-alt);
                display: flex;
                flex-direction: column;
            }

            .meta-pill,
            .signal {
                display: inline-flex;
                align-items: center;
                min-width: 0;
                max-width: 100%;
                padding: 6px 10px;
                border: 1px solid var(--border);
                border-left-width: 2px;
                border-radius: 999px;
                background-color: #120e18;
                color: var(--text-secondary);
                font-family: var(--font-mono);
                font-size: 11px;
                letter-spacing: 0.1em;
                text-transform: uppercase;
                white-space: nowrap;
                overflow: hidden;
                text-overflow: ellipsis;
            }

            .meta-pill.idle,
            .signal.idle {
                border-left-color: var(--border-strong);
            }

            .meta-pill.busy,
            .signal.busy {
                border-color: #383156;
                border-left-color: var(--busy);
                background-color: #141326;
                color: #ddd8ff;
            }

            .meta-pill.ready,
            .signal.ready {
                border-color: #33402a;
                border-left-color: var(--lime);
                background-color: #11160f;
                color: #dfe8cf;
            }

            .section-label {
                display: block;
                margin-bottom: 10px;
                color: var(--text-muted);
                font-family: var(--font-mono);
                font-size: 11px;
                letter-spacing: 0.18em;
                text-transform: uppercase;
            }

            .output-header {
                display: flex;
                align-items: flex-start;
                justify-content: space-between;
                gap: 16px;
                margin-bottom: 16px;
                padding-bottom: 14px;
                border-bottom: 1px solid var(--border);
            }

            .output-controls {
                display: flex;
                align-items: center;
                gap: 8px;
                flex-wrap: wrap;
            }

            .output-title {
                color: var(--text-primary);
                font-size: 20px;
                font-weight: 650;
                letter-spacing: 0.02em;
            }

            textarea {
                width: 100%;
                min-height: 126px;
                padding: 14px 15px;
                border: 1px solid var(--border);
                border-radius: 12px;
                background-color: var(--field-bg);
                color: var(--text-primary);
                font-family: var(--font-sans);
                font-size: 15px;
                line-height: 1.6;
                resize: vertical;
                outline: none;
                transition: background-color 0.14s ease, border-color 0.14s ease, color 0.14s ease;
            }

            textarea:hover {
                border-color: var(--border-strong);
            }

            textarea:focus,
            textarea:focus-visible {
                border-color: var(--accent);
                background-color: var(--field-bg-active);
            }

            textarea::placeholder {
                color: var(--placeholder);
            }

            .drop-zone {
                margin-top: 16px;
                min-height: 76px;
                padding: 14px 16px;
                border: 1px dashed var(--border-strong);
                border-radius: 12px;
                background-color: var(--field-bg);
                cursor: pointer;
                transition: background-color 0.14s ease, border-color 0.14s ease, color 0.14s ease;
            }

            .drop-zone:hover,
            .drop-zone.drag-over {
                border-color: var(--accent);
                background-color: #15101d;
            }

            .drop-zone.busy {
                border-color: #4c4269;
                background-color: #140f1c;
                cursor: progress;
            }

            .drop-inline {
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 12px;
                min-height: 46px;
            }

            .drop-main {
                display: flex;
                align-items: center;
                gap: 12px;
                min-width: 0;
                flex: 1;
            }

            .drop-message {
                min-width: 0;
                color: var(--text-secondary);
                font-size: 14px;
                line-height: 1.5;
                overflow: hidden;
                text-overflow: ellipsis;
                white-space: nowrap;
            }

            .drop-badge {
                flex-shrink: 0;
                padding: 5px 8px;
                border: 1px solid var(--border);
                border-radius: 999px;
                background-color: #161120;
                color: var(--text-muted);
                font-family: var(--font-mono);
                font-size: 10px;
                letter-spacing: 0.14em;
                text-transform: uppercase;
            }

            .filename {
                display: inline;
                margin-left: 8px;
                color: var(--text-primary);
                font-family: var(--font-mono);
                font-size: 12px;
                font-weight: 600;
            }

            .actions {
                margin-top: 16px;
                padding-top: 18px;
                width: 100%;
                border-top: 1px solid var(--border);
            }

            button {
                width: 100%;
                padding: 13px 16px;
                border: 1px solid #4b3a61;
                border-radius: 12px;
                background-color: #2a1b41;
                color: #f3eeff;
                font-family: var(--font-sans);
                font-size: 14px;
                font-weight: 650;
                letter-spacing: 0.03em;
                cursor: pointer;
                transition: background-color 0.14s ease, border-color 0.14s ease, color 0.14s ease;
            }

            button:hover:not(:disabled) {
                border-color: var(--accent-hover);
                background-color: #32204d;
                color: #ffffff;
            }

            button:focus-visible {
                border-color: var(--accent-hover);
                outline: none;
            }

            .button-processing {
                border-color: #58467b;
                background-color: #251a39;
                color: #ddd5ef;
                cursor: wait;
                pointer-events: none;
            }

            .secondary-button {
                width: auto;
                border-color: var(--border);
                background-color: #17121f;
                color: var(--text-secondary);
                font-family: var(--font-mono);
                font-size: 12px;
                letter-spacing: 0.12em;
                text-transform: uppercase;
            }

            .secondary-button:hover:not(:disabled) {
                border-color: var(--border-strong);
                background-color: #1c1626;
                color: var(--text-primary);
            }

            button:disabled {
                opacity: 0.65;
                cursor: not-allowed;
            }

            .copy-button {
                padding: 8px 12px;
                border-radius: 10px;
                line-height: 1;
                flex-shrink: 0;
            }

            pre {
                min-height: calc(100vh - 240px);
                padding: 18px;
                border: 1px solid var(--border);
                border-radius: 12px;
                background-color: var(--console-bg);
                color: #e1dbef;
                font-family: var(--font-mono);
                font-size: 15px;
                line-height: 1.7;
                white-space: pre-wrap;
                word-break: break-word;
                overflow: auto;
                flex: 1;
            }

            .reasoning-toggle {
                width: auto;
                align-self: flex-start;
                margin-top: 12px;
                padding: 10px 14px;
                border-radius: 10px;
            }

            .reasoning-toggle[aria-expanded="true"] {
                border-color: #5f47a2;
                background-color: var(--accent-muted);
                color: #f3eeff;
            }

            .reasoning-panel {
                margin-top: 12px;
                padding: 14px;
                border: 1px solid var(--border);
                border-radius: 12px;
                background-color: #110d17;
            }

            .reasoning-label {
                margin-bottom: 10px;
                color: var(--text-muted);
                font-family: var(--font-mono);
                font-size: 11px;
                letter-spacing: 0.16em;
                text-transform: uppercase;
            }

            .reasoning-output {
                white-space: pre-wrap;
                word-break: break-word;
                color: var(--text-secondary);
                font-family: var(--font-mono);
                font-size: 13px;
                line-height: 1.65;
            }

            .robot-loader {
                width: 100%;
                min-height: 100%;
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
                gap: 12px;
                white-space: normal;
                text-align: center;
                color: var(--text-secondary);
            }

            .loader-ring {
                width: 28px;
                height: 28px;
                border: 3px solid #312741;
                border-top-color: var(--accent);
                border-right-color: rgba(163, 230, 53, 0.65);
                border-radius: 50%;
                animation: spin 0.9s linear infinite;
                flex-shrink: 0;
            }

            .loader-text {
                color: var(--text-muted);
                font-family: var(--font-mono);
                font-size: 12px;
                letter-spacing: 0.12em;
                text-transform: uppercase;
            }

            pre.output-loading {
                display: flex;
                align-items: center;
                justify-content: center;
                white-space: normal;
            }

            @keyframes spin {
                to {
                    transform: rotate(360deg);
                }
            }

            @media (prefers-reduced-motion: reduce) {
                textarea,
                .drop-zone,
                button {
                    transition: none !important;
                }

                .loader-ring {
                    animation: none;
                }
            }

            @media (max-width: 899px) {
                .topbar-row,
                .output-header,
                .drop-inline {
                    flex-direction: column;
                    align-items: flex-start;
                }

                .topbar-pills,
                .output-controls {
                    width: 100%;
                    justify-content: flex-start;
                }

                pre {
                    min-height: 320px;
                }
            }

            @media (max-width: 700px) {
                .page-shell {
                    width: min(100% - 24px, 1180px);
                    padding: 20px 0 24px;
                }

                .top-bar,
                .composer,
                .output-panel {
                    padding: 16px;
                }
            }
        </style>
    </head>
    <body>
        <div class="page-shell">
            <section class="panel hero top-bar">
                <div class="topbar-row">
                    <div class="brand">
                        <svg class="brand-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">
                            <path d="M12 3l7 3v5c0 5-3 8-7 10-4-2-7-5-7-10V6l7-3z"></path>
                            <path d="M9.5 12.5l1.7 1.7 3.3-4"></path>
                        </svg>
                        <div class="app-title">Dark Signal AI</div>
                    </div>
                    <div class="topbar-pills">
                        <span id="file-status" class="meta-pill idle">No file attached</span>
                        <span class="meta-pill">FastAPI + AI Workflow</span>
                    </div>
                </div>
            </section>

            <div class="content-grid">
                <section class="panel composer">
                    <label class="section-label" for="prompt">Prompt Input</label>
                    <textarea id="prompt" rows="5" placeholder="Ask a question about a file, request a summary, or describe the analysis you want."></textarea>

                    <div id="drop-zone" class="drop-zone"></div>
                    <input type="file" id="file-input" hidden>

                    <div class="actions">
                        <button type="button" id="send-btn">Run Analysis</button>
                    </div>
                </section>

                <section class="panel output-panel">
                    <div class="output-header">
                        <div>
                            <div class="section-label">Response Feed</div>
                            <div class="output-title">Model Output</div>
                        </div>
                        <div class="output-controls">
                            <div id="response-state" class="signal idle">Standing by</div>
                            <button type="button" id="copy-output-btn" class="secondary-button copy-button">Copy</button>
                        </div>
                    </div>
                    <pre id="output">Your answer will appear here.</pre>
                    <button type="button" id="reasoning-toggle" class="secondary-button reasoning-toggle" aria-controls="reasoning-panel" aria-expanded="false" hidden>Show reasoning</button>
                    <div id="reasoning-panel" class="reasoning-panel" hidden>
                        <div class="reasoning-label">Model Reasoning</div>
                        <div id="reasoning-output" class="reasoning-output"></div>
                    </div>
                </section>
            </div>
        </div>

        <script>
document.addEventListener("DOMContentLoaded", () => {
    let uploadedFile = null;
    let isUploading = false;
    let isAnalyzing = false;

    const UPLOAD_TIMEOUT_MS = 60000;

    const dropZone = document.getElementById("drop-zone");
    const fileStatus = document.getElementById("file-status");
    const responseState = document.getElementById("response-state");
    const output = document.getElementById("output");
    const promptInput = document.getElementById("prompt");
    const sendButton = document.getElementById("send-btn");
    const copyButton = document.getElementById("copy-output-btn");
    const fileInput = document.getElementById("file-input");
    const reasoningToggle = document.getElementById("reasoning-toggle");
    const reasoningPanel = document.getElementById("reasoning-panel");
    const reasoningOutput = document.getElementById("reasoning-output");

    function escapeHtml(text) {
        return text.replaceAll("&", "&amp;").replaceAll("<", "&lt;").replaceAll(">", "&gt;");
    }

    function updateSendButton() {
        sendButton.classList.toggle("button-processing", isAnalyzing);

        if (isAnalyzing) {
            sendButton.disabled = true;
            sendButton.textContent = "Analyzing...";
            return;
        }

        if (isUploading) {
            sendButton.disabled = true;
            sendButton.textContent = "Uploading file...";
            return;
        }

        sendButton.disabled = false;
        sendButton.textContent = "Run Analysis";
    }

    function updateCopyButton() {
        copyButton.disabled = isAnalyzing;
    }

    function canStartUpload() {
        return !isUploading && !isAnalyzing;
    }

    function renderDropZone() {
        if (isUploading) {
            dropZone.innerHTML = '<div class="drop-inline"><div class="drop-main"><div class="drop-message">Uploading file...</div></div><div class="drop-badge">Please wait</div></div>';
            return;
        }
        if (uploadedFile) {
            dropZone.innerHTML = '<div class="drop-inline"><div class="drop-main"><div class="drop-message">File attached <span class="filename">' + escapeHtml(uploadedFile.name) + '</span></div></div><div class="drop-badge">Attached</div></div>';
            return;
        }
        dropZone.innerHTML = '<div class="drop-inline"><div class="drop-main"><div class="drop-message">Drop file or click to attach</div></div><div class="drop-badge">Any format</div></div>';
    }

    function setFileStatus(text, stateClass) {
        fileStatus.textContent = text;
        fileStatus.className = "meta-pill " + stateClass;
    }

    function setResponseState(text, stateClass) {
        responseState.textContent = text;
        responseState.className = "signal " + stateClass;
    }

    async function readJsonResponse(res) {
        const contentType = (res.headers.get("content-type") || "").toLowerCase();

        if (contentType.includes("application/json")) {
            let data;
            try {
                data = await res.json();
            } catch (error) {
                throw new Error("Server returned invalid JSON.");
            }
            if (!res.ok) {
                throw new Error(data.error || data.detail || `HTTP ${res.status}`);
            }
            return data;
        }

        const text = (await res.text()).trim();
        if (!res.ok) {
            if (text && !text.startsWith("<!DOCTYPE") && !text.startsWith("<html")) {
                throw new Error(text);
            }
            throw new Error(`HTTP ${res.status}`);
        }

        throw new Error("Server returned an unexpected response.");
    }

    async function fetchWithTimeout(url, options, timeoutMs) {
        const controller = new AbortController();
        const timeoutId = setTimeout(() => controller.abort(), timeoutMs);

        try {
            return await fetch(url, { ...options, signal: controller.signal });
        } catch (error) {
            if (error instanceof DOMException && error.name === "AbortError") {
                throw new Error("Request timed out. Please try again.");
            }
            throw error;
        } finally {
            clearTimeout(timeoutId);
        }
    }

    function hasReasoning(reasoning) {
        return typeof reasoning === "string" && reasoning.trim().length > 0;
    }

    function resetReasoningUI() {
        reasoningOutput.textContent = "";
        reasoningPanel.hidden = true;
        reasoningToggle.hidden = true;
        reasoningToggle.textContent = "Show reasoning";
        reasoningToggle.setAttribute("aria-expanded", "false");
    }

    function setReasoning(reasoning) {
        if (!hasReasoning(reasoning)) {
            resetReasoningUI();
            return;
        }

        reasoningOutput.textContent = reasoning;
        reasoningPanel.hidden = true;
        reasoningToggle.hidden = false;
        reasoningToggle.textContent = "Show reasoning";
        reasoningToggle.setAttribute("aria-expanded", "false");
    }

    let copyResetTimer = null;

    function setCopyButtonLabel(text) {
        copyButton.textContent = text;
        if (copyResetTimer) {
            clearTimeout(copyResetTimer);
        }
        if (text !== "Copy") {
            copyResetTimer = setTimeout(() => {
                copyButton.textContent = "Copy";
                copyResetTimer = null;
            }, 1200);
        }
    }

    async function handleFile(file) {
        if (!canStartUpload()) {
            return;
        }
        isUploading = true;
        uploadedFile = null;
        updateSendButton();
        renderDropZone();
        dropZone.classList.add("busy");
        setFileStatus("Uploading file", "busy");
        try {
            const form = new FormData();
            form.append("file", file);
            const res = await fetchWithTimeout("/upload", { method: "POST", body: form }, UPLOAD_TIMEOUT_MS);
            const data = await readJsonResponse(res);
            uploadedFile = { id: data.file_id, name: data.filename };
            setFileStatus("Attached: " + data.filename, "ready");
        } catch (err) {
            uploadedFile = null;
            output.textContent = err instanceof Error ? err.message : "Upload failed.";
            resetReasoningUI();
            setFileStatus("Upload failed", "idle");
        } finally {
            isUploading = false;
            updateSendButton();
            renderDropZone();
            dropZone.classList.remove("busy");
        }
    }

    async function sendPrompt() {
        const prompt = promptInput.value;
        if (isUploading) {
            output.textContent = "Please wait for the file upload to finish.";
            setResponseState("Upload in progress", "busy");
            return;
        }
        if (!prompt.trim()) {
            resetReasoningUI();
            output.textContent = "Please enter a prompt first.";
            setResponseState("No prompt entered", "idle");
            return;
        }
        resetReasoningUI();
        isAnalyzing = true;
        updateSendButton();
        updateCopyButton();
        output.textContent = "";
        output.classList.add("output-loading");
        output.innerHTML = '<div class="robot-loader" role="status" aria-live="polite"><span class="loader-ring" aria-hidden="true"></span><span class="loader-text">Analyzing...</span></div>';
        setResponseState("Analyzing", "busy");
        try {
            const body = { prompt };
            if (uploadedFile) {
                body.file_id = uploadedFile.id;
            }
            const res = await fetch("/ask", {
                method: "POST",
                headers: {"Content-Type": "application/json"},
                body: JSON.stringify(body)
            });
            const data = await readJsonResponse(res);
            output.classList.remove("output-loading");
            output.textContent = data.answer || data.error || "No response.";
            if (hasReasoning(data.reasoning)) {
                setReasoning(data.reasoning);
            }
            setResponseState("Response ready", "ready");
        } catch (error) {
            output.classList.remove("output-loading");
            output.textContent = error instanceof Error ? error.message : "Request failed. Please try again.";
            resetReasoningUI();
            setResponseState("Request failed", "idle");
        } finally {
            isAnalyzing = false;
            updateSendButton();
            updateCopyButton();
        }
    }

    renderDropZone();
    updateSendButton();
    updateCopyButton();

    dropZone.addEventListener("click", () => {
        if (canStartUpload()) {
            fileInput.click();
        }
    });
    dropZone.addEventListener("dragover", e => {
        e.preventDefault();
        if (canStartUpload()) {
            dropZone.classList.add("drag-over");
        }
    });
    dropZone.addEventListener("dragleave", () => dropZone.classList.remove("drag-over"));
    dropZone.addEventListener("drop", e => {
        e.preventDefault();
        dropZone.classList.remove("drag-over");
        if (canStartUpload() && e.dataTransfer.files.length) handleFile(e.dataTransfer.files[0]);
    });
    fileInput.addEventListener("change", e => {
        const selectedFile = e.target.files && e.target.files[0];
        fileInput.value = "";
        if (canStartUpload() && selectedFile) handleFile(selectedFile);
    });
    sendButton.addEventListener("click", sendPrompt);
    reasoningToggle.addEventListener("click", () => {
        const shouldShow = reasoningPanel.hidden;
        reasoningPanel.hidden = !shouldShow;
        reasoningToggle.textContent = shouldShow ? "Hide reasoning" : "Show reasoning";
        reasoningToggle.setAttribute("aria-expanded", shouldShow ? "true" : "false");
    });
    copyButton.addEventListener("click", async () => {
        if (output.classList.contains("output-loading")) {
            setCopyButtonLabel("Wait");
            return;
        }
        const text = output.textContent.trim();
        if (!text || text === "Your answer will appear here.") {
            setCopyButtonLabel("Empty");
            return;
        }
        try {
            await navigator.clipboard.writeText(text);
            setCopyButtonLabel("Copied");
        } catch (error) {
            setCopyButtonLabel("Failed");
        }
    });
});
</script>
    </body>
    </html>
    """
